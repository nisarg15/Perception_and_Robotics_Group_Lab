import os
import numpy as np
import torch
from torch.utils.data import Dataset
import cv2
import pickle   
import math
import matplotlib.pyplot as plt
import time
from io import BytesIO
import shutil
import gc
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import openpyxl

# Create a new workbook and select the active sheet


series = []
indexs = []
total_difference = []
total_angle = []
total_points = []
total_sum = []
totle_angle_sum = []






class NormalFlowSupervisedDataset(Dataset):
    def __init__(self, sequence_paths, network="NormalFlow", target="NormalFlow", transform=None, ):
        # List of sequences that will be used
        self.sequence_paths = sequence_paths

        self.transform = transform
        self.network = network
        self.target = target
        

        # Find out the total number of samples
        self.num_samples = 0
        self.sequence_start_index = np.zeros((len(self.sequence_paths),), dtype=np.int64)
        for i, sequence_path in enumerate(self.sequence_paths):
            nf_path = os.path.join(sequence_path, 'dataset_normal_flow.npz')
            nf_data = np.load(nf_path)

            self.sequence_start_index[i] = self.num_samples
            self.num_samples += nf_data['num_of_samples']

        self.opened_in_out_pairs = None

    def __len__(self):
        return self.num_samples

    def __getitem__(self, idx):
        # Pre-open all the npz pairs to get a 50% boost in single thread 25% in multithread
        # Use this hack to pre-open them to because NumPy does not support sharing npz objects
        # across processes
        if self.opened_in_out_pairs is None:
            self.opened_in_out_pairs = {}
            for sequence_path in self.sequence_paths:
                sequence = os.path.split(sequence_path)[-1]
                nf_path = os.path.join(sequence_path, 'dataset_normal_flow.npz')
                self.opened_in_out_pairs[sequence] = np.load(nf_path)

        seq_index = np.searchsorted(self.sequence_start_index, idx, side='right') - 1
        sequence_path = self.sequence_paths[seq_index]
        sequence = os.path.split(sequence_path)[-1]

        in_out_pairs = self.opened_in_out_pairs[sequence]
        

        frame_idx = in_out_pairs['index_to_frame_id'][idx - self.sequence_start_index[seq_index]]
        rjust_frame_idx = str(frame_idx).rjust(10, '0')
       
        X_pca  = in_out_pairs[f'normal_flow_pca_{rjust_frame_idx}']
        X_ec   = in_out_pairs[f'normal_flow_event_count_{rjust_frame_idx}']
        X_ts   = in_out_pairs[f'normal_flow_event_timestamp_{rjust_frame_idx}']

        if self.network == "NormalFlow":
            X = np.dstack((X_pca, X_ec))     
        elif self.network == "OpticalFlow":
            X = np.dstack((X_ec, X_ts))
        else:
            raise ValueError(f"Incorrect Network Name {self.network}, Use NormalFlow or OpticalFlow")
            
        if self.target == "NormalFlow":
            Y      = in_out_pairs[f'normal_flow_gt_{rjust_frame_idx}']
            Y_mask = in_out_pairs[f'normal_flow_gt_mask_{rjust_frame_idx}']
        elif self.target == "OpticalFlow":
            Y      = in_out_pairs[f'optical_flow_gt_{rjust_frame_idx}']
            Y_mask = in_out_pairs[f'optical_flow_gt_mask_{rjust_frame_idx}']
        elif self.target == "BGR":
            Y      = in_out_pairs[f'bgr_image_{rjust_frame_idx}']
            Y_mask = in_out_pairs[f'bgr_image_mask_{rjust_frame_idx}']
        else:
            raise ValueError(f"Incorrect Target Name {self.target}, Use NormalFlow or OpticalFlow")

        X      = torch.tensor(np.nan_to_num(X)).permute(2, 0, 1)
        Y      = torch.tensor(np.nan_to_num(Y)).permute(2, 0, 1)
        Y_mask = torch.tensor(Y_mask)

        if self.transform:
            X = self.transform(X)
            Y = self.transform(Y)
            Y_mask = self.transform(Y_mask)
       

        return X, Y, Y_mask
    
        
def visualize_optical_flow(flowin):

    flow=np.ma.array(flowin, mask=np.isnan(flowin))
    theta = np.mod(np.arctan2(flow[:, :, 0], flow[:, :, 1]) + 2*np.pi, 2*np.pi)
    flow_norms = np.linalg.norm(flow, axis=2)
    flow_norms[flow_norms > 50] = 0
    flow_norms_normalized = flow_norms / np.max(np.ma.array(flow_norms, mask=np.isnan(flow_norms)))
    empty_image = np.zeros((480, 640, 3))
    empty_image[:,:,0] = flow_norms_normalized
    flow_hsv = np.zeros((flow.shape[0], flow.shape[1], 3), dtype=np.uint8)
    flow_hsv[:, :, 0] = 179 * theta / (2*np.pi)
    flow_hsv[:, :, 1] = 255 * (flow_norms_normalized >0 )
    flow_hsv[:, :, 2] = 255 * (flow_norms_normalized > 0) 
    flow_hsv[np.logical_and(np.isnan(theta), np.isnan(flow_norms_normalized)), :] = 0

    return flow_hsv,empty_image
  
def visualize_optical_flow_A(flowin_A):
    
    flow=np.ma.array(flowin_A, mask=np.isnan(flowin_A))
    theta = np.mod(np.arctan2(flow[:, :, 0], flow[:, :, 1]) + 2*np.pi, 2*np.pi)
    flow_norms = np.linalg.norm(flow, axis=2)
    flow_norms[flow_norms > 50] = 0
    flow_norms_normalized =flow_norms / np.max(np.ma.array(flow_norms, mask=np.isnan(flow_norms)))
    empty_image = np.zeros((480, 640, 3))
    empty_image[:,:,1] = flow_norms_normalized
    flow_hsv = np.zeros((flow.shape[0], flow.shape[1], 3), dtype=np.uint8)
    flow_hsv[:, :, 0] = 179 * theta / (2*np.pi)
    flow_hsv[:, :, 1] = 255 * (flow_norms_normalized >0)
    flow_hsv[:, :, 2] = 255 * (flow_norms_normalized > 0) 
    flow_hsv[np.logical_and(np.isnan(theta), np.isnan(flow_norms_normalized)), :] = 0

    return flow_hsv,empty_image


        
        
def visualize_train_layer(train_layer):
    nf = train_layer[:, :, 0:2]
    ec = train_layer[:, :, 2]
    return visualize_target_layer(ec, nf)

def visualize_target_layer(ec, nf, reconstructed=None):
    
    ec = (128*(ec > 0)).astype(np.uint8)
    ogxx, ogyy, __ = np.nonzero(nf)
    oguuvv = nf[ogxx, ogyy,:]
    oguu, ogvv = oguuvv[:,0], oguuvv[:,1]
    pixel_array = np.dstack((ogxx,ogyy,ogvv,oguu))
    height, width = 480, 640
    image = np.zeros((height, width, 2), dtype=np.float32)

    x, y, r, g = pixel_array[0].T
    x, y = x.astype(int), y.astype(int)
    image[x, y] = np.column_stack((r, g))

    flow_hsv,norm_img_2 = visualize_optical_flow(image)
    flow_bgr = cv2.cvtColor(flow_hsv, cv2.COLOR_HSV2BGR)
    ec_rgb = np.dstack((ec, ec, ec)) 


    return ec_rgb,flow_bgr,norm_img_2,ogyy, ogxx, oguu, ogvv

def visualize_target_layer_nf(ec,xy,nf,reconstructed=None):
    
    ec = (128*(ec > 0)).astype(np.uint8)
    st  =time.time()
    xy = np.array(xy)
    nf = np.array(nf)
    ogxx_a, ogyy_a = xy[:,1], xy[:,0]
    oguu_a, ogvv_a = nf[:,0],nf[:,1]
    pixel_array_a = np.dstack((ogxx_a,ogyy_a,ogvv_a,oguu_a))
    height, width = 480, 640
    image_a = np.zeros((height, width, 2), dtype=np.float32)

    x, y, r, g = pixel_array_a[0].T
    x, y = x.astype(int), y.astype(int)
    image_a[x, y] = np.column_stack((r, g))


    flow_hsv_a,norm_img = visualize_optical_flow_A(image_a)
    flow_bgr_a = cv2.cvtColor(flow_hsv_a, cv2.COLOR_HSV2BGR)
    ec_rgb_a = np.dstack((ec, ec, ec))

    return ec_rgb_a,flow_bgr_a,norm_img


def visualize_pytorch_layers(train_layer, gtnf,xy,nf , gtnf_mask, reconstructed=None):
    train_layer = train_layer[0, :, :].permute(1,2,0).numpy()
    gtnf        = gtnf       [0, :, :].permute(1,2,0).numpy()        
    gtnf_mask   = gtnf_mask  [0, :, :].numpy()

    if reconstructed is not None:
        reconstructed = reconstructed[0, :, :].permute(1,2,0).numpy

    assert train_layer.dtype == np.float32
    assert gtnf.dtype == np.float32
    assert gtnf_mask.dtype == bool
    assert gtnf.max() <= 640
    assert gtnf.min() >= -640

    gtnf[gtnf_mask==False, :] = 0

    vis_gt,img,norm_img,ogyy, ogxx, oguu, ogvv = visualize_target_layer(train_layer[:, :, 2], gtnf, reconstructed=reconstructed)
    vis_train,img_a,norm_img_2 = visualize_target_layer_nf(train_layer[:, :, 2], xy,nf,reconstructed=reconstructed)

    return vis_train, vis_gt ,img,img_a,norm_img,norm_img_2,ogyy, ogxx, oguu, ogvv

def visualize_layers(t_file, d):
    l = DataLoader(d, batch_size=1, shuffle=False, num_workers=10)
    series = np.array(t_file)
    indices = np.abs(series[:, None] - t_file).argmin(axis=0)

    return indices.tolist(), l

if __name__ == '__main__':
    import glob
    from tqdm import tqdm
    from torch.utils.data import DataLoader  
    folder = "scene15_dyn_test_03_000000"   
    pc = "enigma"   

    sequence_paths = sorted(list(glob.glob(f'/home/{pc}/prg/DeepFit/tutorial/*')))
    assert len(sequence_paths) > 0
    sequence_paths = [f'/home/{pc}/prg/DeepFit/tutorial/{folder}/']
    t_file = f"/home/{pc}/prg/DeepFit/tutorial/{folder}/dataset_normal_flow/t.npy"
    d = NormalFlowSupervisedDataset(sequence_paths)
    t_file = np.load(t_file)
    t_file = t_file - t_file[0]
    


    indexs,l = visualize_layers(t_file,d)
    indexs = np.array(indexs)
    indexs = indexs+1
    counter = 1
    totle_index_len = len(indexs)
    frame_width = 2560
    frame_height = 1440
    output_video_path = f'/home/{pc}/prg/DeepFit/tutorial/{folder}/output_video_2.avi'
    fourcc = cv2.VideoWriter_fourcc(*'XVID')  # Codec for AVI format
    out = cv2.VideoWriter(output_video_path, fourcc, 10, (frame_width, frame_height)) 
    s = time.time()

    hsv_mag_img_path = "/home/enigma/prg/DeepFit/tutorial/Lab Images/MAG_VALUES.png"
    hsv_ang_img_path = "/home/enigma/prg/DeepFit/tutorial/Lab Images/ANG_VALUES.png"
    hsv_mag_img = cv2.imread(hsv_mag_img_path)
    hsv_ang_img = cv2.imread(hsv_ang_img_path)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for train_layer, gtnf, gtnf_mask in l:
        totle_mean_mag = []
        totle_median_mag = []
        totle_std_mag = []

        totle_mean_ang = []
        totle_median_ang = []
        totle_std_ang = []

        totle_max_mag = []
        totle_min_mag = []

        totle_max_ang = []
        totle_min_ang = []

        angle_1 = []
        angle_2 = []
        angle_diff = []

        if counter == totle_index_len:
            break

        if indexs[1]-indexs[0]==0:
        
            xy = f"/home/{pc}/prg/DeepFit/tutorial/{folder}/nf/{folder}_{indexs[0]}.xyz"
            nf = f"/home/{pc}/prg/DeepFit/tutorial/{folder}/nf/{folder}_{indexs[0]}.nf"

            print(xy)  
            with open(xy, 'rb') as file:
                xy_data = pickle.load(file)
            with open(nf, 'rb') as file:
                nf_data = pickle.load(file)
            indexs = indexs[1:]

           
            
        else:
            cc = indexs[0]
             
            for i in range(indexs[1]-indexs[0]):
                
                xy = f"/home/{pc}/prg/DeepFit/tutorial/{folder}/nf/{folder}_{cc}.xyz"
                nf = f"/home/{pc}/prg/DeepFit/tutorial/{folder}/nf/{folder}_{cc}.nf"
                print("HIIIIII")
                print(xy)  

                with open(xy, 'rb') as file:
                    xy_data = pickle.load(file)
                with open(nf, 'rb') as file:
                    nf_data = pickle.load(file)
            indexs = indexs[1:]

                
        vis_train, vis_gt,img,img_a,norm_img,norm_img_2,ogyy, ogxx, oguu, ogvv = visualize_pytorch_layers(train_layer, gtnf,xy_data,nf_data ,gtnf_mask)
        combined_image = cv2.add(norm_img, norm_img_2)
        green_mask = (norm_img[:, :, 0] > 0) & (norm_img_2[:, :, 1] > 0)

# Create a red image with the same size as the input images
        result_image = np.zeros_like(norm_img_2)
        result_image[green_mask] = [0, 0, 255] 

        lower_red = np.array([0, 0, 100])
        upper_red = np.array([100, 100, 255])

        # Create a mask where red pixels are present
        red_mask = cv2.inRange(result_image, lower_red, upper_red)

        # Find coordinates of red pixels
        red_pixel_coordinates = np.argwhere(red_mask > 0)

        red_pixel_coordinates = red_pixel_coordinates[:, ::-1]
        red_pixel_coordinates = np.array(red_pixel_coordinates)
        
        
        gt_xy_data = np.column_stack((ogyy, ogxx))
        gt_nf_data = np.column_stack((oguu, ogvv))
        xy_data = np.array(xy_data)
        
        xy_data_tuple = [tuple(inner_list) for inner_list in xy_data]
        red_pixel_coordinates_tuple = [tuple(inner_list) for inner_list in red_pixel_coordinates]

        my_dict = dict(zip(xy_data_tuple, nf_data))
        values = np.array([my_dict[key] for key in red_pixel_coordinates_tuple if key in my_dict])

        selected_values_nf = values
        selected_values_xy = red_pixel_coordinates

        selected_values_nf_new = selected_values_xy + selected_values_nf

        magnitudes_array = np.sqrt(np.sum(selected_values_nf**2, axis=1))
        magnitudes_array[magnitudes_array > 50] = 0
        x_values = np.arange(len(magnitudes_array))


        
        # height, width = 640, 480  # Replace with your actual dimensions

        # # Initialize the boolean array
        # red_pixel_coordinates_array = np.zeros((height, width), dtype=bool)

        # # Set the red pixels to True in the boolean array
        # red_pixel_coordinates_array[red_pixel_coordinates[:, 0], red_pixel_coordinates[:, 1]] = True
        # keep = red_pixel_coordinates_array[gt_xy_data]


        # print(np.shape(values_gt))

        # red_pixel_coordinates_array = np.zeros((height, width, 2))
        # red_pixel_coordinates_array.at(red_pixel_coordinates[:, 0], red_pixel_coordinates[:, 1]) = 1
        # keep = red_pixel_coordinates_array[gt_xy_data]
        # values_gt = gt_nf_data[keep]
    
        gt_xy_data_tuple = [tuple(inner_list) for inner_list in gt_xy_data]

        my_dict_gt = dict(zip(gt_xy_data_tuple, gt_nf_data))
        values_gt = np.array([my_dict_gt[key] for key in red_pixel_coordinates_tuple if key in my_dict_gt])
   

        



        gt_selected_values = values_gt
        gt_selected_values_xy = red_pixel_coordinates
        gt_selected_values_new = gt_selected_values_xy + gt_selected_values

        gt_magnitudes_array = np.sqrt(np.sum(gt_selected_values**2, axis=1))
        
        gt_magnitudes_array[gt_magnitudes_array > 50] = 0
        combined_image[result_image[:, :, 2] > 0, :] = 0
        combined_image[result_image[:, :, 2] > 0, 2] = result_image[result_image[:, :, 2] > 0, 2]

        difference = np.array(magnitudes_array-gt_magnitudes_array)
        difference =  np.abs(difference)
 
        x1 = selected_values_xy[:, 0]
        y1 = selected_values_xy[:, 1]

        x2_v1 = selected_values_nf_new[:, 0]
        y2_v1 = selected_values_nf_new[:, 1]

        x2_v2 = gt_selected_values_new[:, 0]
        y2_v2 = gt_selected_values_new[:, 1]

        # Calculate angles
        angle_2 = np.degrees(np.arctan2(y2_v1 - y1, x2_v1 - x1))
        angle_1 = np.degrees(np.arctan2(y2_v2 - y1, x2_v2 - x1))

        # Calculate angle differences
        start_a = np.column_stack((x1, y1))
        end_a = np.column_stack((x2_v1, y2_v1))
        start_b = np.column_stack((x1, y1))
        end_b = np.column_stack((x2_v2, y2_v2))

        a = end_a - start_a
        b = end_b - start_b

        dot_product = np.sum(a * b, axis=1)
        magnitude_a = np.linalg.norm(a, axis=1)
        magnitude_b = np.linalg.norm(b, axis=1)
       


        angle_diff = np.degrees(np.arccos(np.divide(dot_product, magnitude_a * magnitude_b, out=np.zeros_like(dot_product), where=magnitude_a * magnitude_b != 0)))
        angle_diff[np.isnan(angle_diff)] = 0

 
        multipler = 179/180
        hue_array = angle_diff * multipler
        hue_array_1 = np.array([(difference / max(difference)) * 179])

        

        # Create an empty HSV image
        height, width = 480, 640  # Set the dimensions of the image
        hsv_image = np.zeros((height, width, 3), dtype=np.uint8)

        # Fill the HSV image with hue, saturation, and value
        x_coords, y_coords = zip(*gt_selected_values_xy)

        # Perform vectorized assignment
        hsv_image[y_coords, x_coords, 0] = hue_array / 2  # Hue
        hsv_image[y_coords, x_coords, 1] = 255          # Saturation
        hsv_image[y_coords, x_coords, 2] = 255   
        

        bgr_image = cv2.cvtColor(hsv_image, cv2.COLOR_HSV2BGR)



        hsv_image_1 = np.zeros((height, width, 3), dtype=np.uint8)

        # Fill the HSV image with hue, saturation, and value
        hsv_image_1[y_coords, x_coords, 0] = hue_array_1 / 2  # Hue
        hsv_image_1[y_coords, x_coords, 1] = 255          # Saturation
        hsv_image_1[y_coords, x_coords, 2] = 255   

        bgr_image_1 = cv2.cvtColor(hsv_image_1, cv2.COLOR_HSV2BGR)
        
        mag_scale = 5
        color  = (0,255,0)
        st = time.time()
        
        xx = selected_values_xy[:,0]
        yy = selected_values_xy[:,1]
        flow_x = selected_values_nf[:,0]
        flow_y = selected_values_nf[:,1]



        for x, y, u, v in zip(xx, yy, flow_x, flow_y):
            if np.sqrt((x+mag_scale*u- x)**2 + (y+mag_scale*v - y)**2) > 50 or np.isnan(x+mag_scale*u) or np.isnan(y+mag_scale*v) :
                continue

            cv2.arrowedLine(vis_train,
                        (int(x), int(y)),
                        (int(x+mag_scale*u), int(y+mag_scale*v)),
                        color,
                        tipLength=0.2)
       

        xx = gt_selected_values_xy[:,0]
        yy = gt_selected_values_xy[:,1]
        flow_x = gt_selected_values[:,0]
        flow_y = gt_selected_values[:,1]


        for x, y, u, v in zip(xx, yy, flow_x, flow_y):
            if np.sqrt((x+mag_scale*u- x)**2 + (y+mag_scale*v - y)**2) > 50 or np.isnan(x+mag_scale*u) or np.isnan(y+mag_scale*v) :
                continue

            cv2.arrowedLine(vis_gt,
                        (int(x), int(y)),
                        (int(x+mag_scale*u), int(y+mag_scale*v)),
                        color,
                        tipLength=0.2)
            
    
        result_array = [] 
        angle_1 = np.array(angle_1)
        angle_diff = np.array(angle_diff)

        mask = angle_1 < 0
        result_array = np.where(mask, -angle_diff, angle_diff)


        sorted_indices = np.argsort(angle_1)
        sorted_another_array = result_array[sorted_indices]

        sorted_indices_1 = np.argsort(gt_magnitudes_array)
        sorted_another_array_1 = difference[sorted_indices_1]

        totle_max_mag.append(np.max(difference))
        totle_min_mag.append(np.min(difference))
        totle_mean_mag.append(np.mean(difference))
        totle_median_mag.append(np.median(difference))
        totle_std_mag.append(np.std(difference))


        totle_max_ang.append(np.max(angle_diff))
   
        totle_min_ang.append(np.min(angle_diff))


        if np.isnan(np.mean(angle_diff)):
            totle_mean_ang.append(0)
        else:
            totle_mean_ang.append(np.mean(angle_diff))
    
        # print("MEDIAN", np.median(angle_diff))
        if np.isnan(np.median(angle_diff)):
            totle_median_ang.append(0)
        else:
            totle_median_ang.append(np.median(angle_diff))
 
        # print("STD", np.std(angle_diff))
        if np.isnan(np.std(angle_diff)):
            totle_std_ang.append(0)
        else:
            totle_std_ang.append(np.std(angle_diff))


        totle_mean_angs = np.nan_to_num(totle_mean_ang, nan=0.0)
        totle_median_angs = np.nan_to_num(totle_median_ang, nan=0.0)
        totle_std_angs = np.nan_to_num(totle_mean_ang, nan=0.0)

      
        
    
        totle_max_angss = np.round(totle_max_ang, decimals=2)
        totle_min_angss = np.round(totle_min_ang, decimals=2)
        totle_mean_angss = np.round(totle_mean_angs, decimals=2)
        totle_std_angss = np.round(totle_std_angs, decimals=2)

        totle_max_magss = np.round(totle_max_mag, decimals=2)
        totle_min_magss = np.round(totle_min_mag, decimals=2)
        totle_mean_magss = np.round(totle_mean_mag, decimals=2)
        totle_std_magss = np.round(totle_std_mag, decimals=2)


        plt.figure(figsize=(6, 4)) 
        plt.plot(x_values, np.sort(angle_diff), label='Ang_Difference', color='blue')
        plt.title(("Max" , totle_max_angss,"Min",totle_min_angss, "Mean",totle_mean_angss, "STD", totle_std_angss ))
        plt.xlabel('X')
        plt.ylabel('ANgles')
        plt.legend()
        plt.savefig('Ang_Difference.png', format='png')
        # plt.pause(0.1)
        # plt.close()
        Ang_Difference = cv2.imread('Ang_Difference.png')
        os.remove('Ang_Difference.png')


        plt.figure(figsize=(6, 4)) 
        plt.plot(x_values, np.sort(difference), label='Mag_Difference', color='black')
        plt.title(("Max" , totle_max_magss,"Min",totle_min_magss, "Mean",totle_mean_magss, "STD", totle_std_magss ))
        plt.xlabel('X')
        plt.ylabel('MAG')
        plt.legend()
        plt.savefig('Mag_Difference.png', format='png')
        # plt.pause(0.1)
        # plt.close() 
        Mag_Difference = cv2.imread('Mag_Difference.png')
        os.remove('Mag_Difference.png')

        plt.figure(figsize=(6, 4)) 
        plt.plot(x_values, sorted_another_array_1, label='Magnitude_Difference', color='blue')
        plt.plot(x_values, np.sort(gt_magnitudes_array), label='GT', color='orange')
        plt.xlabel('X')
        plt.ylabel('Magnitude')
        plt.legend()
        plt.savefig('Magnitude_Difference.png', format='png')
        # plt.pause(0.1)
        # plt.close() 
        Magnitude_Difference = cv2.imread('Magnitude_Difference.png')
        os.remove('Magnitude_Difference.png')
        
    

        plt.figure(figsize=(6, 4)) 
        plt.plot(x_values, sorted_another_array, label='Angle_difference', color='blue')
        plt.plot(x_values, np.sort(angle_1), label='GT', color='orange')
        plt.xlabel('X')
        plt.ylabel('Magnitude')
        plt.legend()
        plt.savefig('Angle_difference.png', format='png')
        # plt.pause(0.1)
        # plt.close()
        gc.collect()    
        Angle_difference = cv2.imread('Angle_difference.png')
        os.remove('Angle_difference.png')

        plt.pause(0.1)
        plt.close("all")

        height1, width1 = img.shape[:2]
        Ang_Difference = cv2.resize(Ang_Difference, (width1, height1))
        Mag_Difference = cv2.resize(Mag_Difference, (width1, height1))
        text = "Ground_Truth"

        # Define the font settings
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_size = 0.7
        font_color = (255, 255, 255)  # White color in BGR
        font_thickness = 2

        # Define the position to start writing text (bottom-left corner)
        text_position = (10, 30)

        # Add the text to the image
        cv2.putText(img, text, text_position, font, font_size, font_color, font_thickness)

        text = "Ground_Truth_NF"
        cv2.putText(vis_gt, text, text_position, font, font_size, font_color, font_thickness)


        plots = np.hstack((img,vis_gt,Ang_Difference,Mag_Difference))

        # Define the font settings
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_size = 0.7
        font_color = (255, 255, 255)  # White color in BGR
        font_thickness = 2

        # Define the position to start writing text (bottom-left corner)
        text_position = (10, 30)
        text = "Cuboid_Method"

        # Add the text to the image
        cv2.putText(img_a, text, text_position, font, font_size, font_color, font_thickness)
        text = "Cuboid_Method_NF"

        # Add the text to the image
        cv2.putText(vis_train, text, text_position, font, font_size, font_color, font_thickness)

        Angle_difference = cv2.resize(Angle_difference, (width1, height1))
        Magnitude_Difference = cv2.resize(Magnitude_Difference, (width1, height1))


        new_photos = np.hstack((img_a,vis_train,Angle_difference,Magnitude_Difference))
  
        height1, width1 = new_photos.shape[:2]
        plots = cv2.resize(plots, (width1, height1))


        text = "Overlapping_Angle_Difference"

        # Define the position to start writing text (bottom-left corner)
        text_position = (10, 30)

        # Add the text to the image
        cv2.putText(bgr_image, text, text_position, font, font_size, font_color, font_thickness)

        text = "Angle_Difference_Values"

        # Add the text to the image
        cv2.putText(hsv_ang_img, text, text_position, font, font_size, (0,0,0), font_thickness)

        text = "Overlapping_Magnitude_Difference"

        # Add the text to the image
        cv2.putText(bgr_image_1, text, text_position, font, font_size, font_color, font_thickness)

        text = "Magnitude_Difference_Values"

        # Add the text to the image
        cv2.putText(hsv_mag_img, text, text_position, font, font_size, (0,0,0), font_thickness)
        photos = np.hstack((bgr_image,hsv_ang_img,bgr_image_1,hsv_mag_img))

        f_image = np.vstack((plots,new_photos,photos))


        counter = counter+1 
        
        total_difference.append(np.mean(difference))
        total_angle.append(np.mean(angle_diff))
        total_points.append(len(red_pixel_coordinates))
        sum_of_mag = (len(red_pixel_coordinates)*np.mean(difference))
        sum_of_ang = (len(red_pixel_coordinates)*np.mean(angle_diff))
        total_sum.append(sum_of_mag)
        totle_angle_sum.append(sum_of_ang)


        out.write(f_image)
        sheet.cell(row = 1 , column = 1,value = "Frames")
        sheet.cell(row = 1 , column = 2,value = "Magnitude_Difference")
        sheet.cell(row = 1 , column = 3,value = "Angle_Difference")
        sheet.cell(row = 1 , column = 4,value = "Number of points")
        sheet.cell(row = 1 , column = 5,value = "Magnitude_Difference_points")
        sheet.cell(row = 1 , column = 6,value = "Angle_Difference_points")
        sheet.cell(row=counter, column=1, value="Time_stops" + str(counter - 1))
        sheet.cell(row=counter, column=2, value=np.mean(difference))
        sheet.cell(row=counter, column=3, value=np.mean(angle_diff))
        sheet.cell(row=counter, column=4, value=len(red_pixel_coordinates))
        sheet.cell(row=counter, column=5, value=(len(red_pixel_coordinates)*np.mean(difference)))
        sheet.cell(row=counter, column=6, value=(len(red_pixel_coordinates)*np.mean(angle_diff)))

    sheet.cell(row = counter+1 , column = 1,value = "Average")
    sheet.cell(row=counter+1, column=2, value=np.mean(total_difference))
    sheet.cell(row=counter+1, column=3, value=np.mean(total_angle))
    sheet.cell(row=counter+1, column=5, value=(sum(total_sum)/sum(total_points)))
    sheet.cell(row=counter+1, column=6, value=(sum(totle_angle_sum)/sum(total_points)))
   
    out.release()
    workbook.save(f'/home/{pc}/prg/DeepFit/tutorial/{folder}/output.xlsx')
    e = time.time()
    print(e-s,"time")
